import sys, io
import os, datetime as dt
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from scipy.optimize import brentq

RS = '\u20b9'

# ── XIRR ─────────────────────────────────────────────────────────────────────
def xirr(cashflows):
    if len(cashflows) < 2:
        return None
    d0 = cashflows[0][0]
    days = [(d - d0).days for d, v in cashflows]
    vals = [v for d, v in cashflows]
    def npv(r):
        return sum(v / (1 + r) ** (d / 365) for v, d in zip(vals, days))
    try:
        return round(brentq(npv, -0.999, 50.0) * 100, 2)
    except Exception:
        return None

# ── STYLE CONSTANTS ───────────────────────────────────────────────────────────
DARK_BLUE = '1F3864'
MID_BLUE  = '2E75B6'
SEC_BLUE  = '4472C4'
LT_BLUE   = 'BDD7EE'
VLT_BLUE  = 'DEEAF1'
GREEN     = '375623'
RED       = 'C00000'
AMBER     = '7F4A00'
WHITE     = 'FFFFFF'
GREY      = 'F2F2F2'

thin = Side(style='thin', color='B8CCE4')
bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

def mkfill(c): return PatternFill('solid', fgColor=c)

def put(ws, r, c, v, bg=None, fc='000000', bold=False, ha='left',
        sz=10, italic=False, merge_to=None, row_h=None):
    cl = ws.cell(row=r, column=c, value=v)
    if bg: cl.fill = mkfill(bg)
    cl.font      = Font(bold=bold, color=fc, size=sz, italic=italic, name='Calibri')
    cl.alignment = Alignment(horizontal=ha, vertical='center')
    cl.border    = bdr
    if merge_to:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=merge_to)
    if row_h:
        ws.row_dimensions[r].height = row_h
    return cl

def sec(ws, r, title):
    put(ws, r, 1, f'  {title}', bg=SEC_BLUE, fc=WHITE, bold=True, sz=11,
        ha='left', merge_to=8, row_h=22)

def pair(ws, r, l1, v1, l2=None, v2=None, fc1='1F3864', fc2='1F3864'):
    bg = VLT_BLUE if r % 2 == 0 else WHITE
    put(ws, r, 1, l1, bg=bg, bold=True, ha='right', sz=10)
    put(ws, r, 2, v1, bg=bg, fc=fc1, bold=True, ha='left', sz=11)
    put(ws, r, 3, '',  bg=bg)
    put(ws, r, 4, l2 or '', bg=bg, bold=True, ha='right', sz=10)
    put(ws, r, 5, v2 or '', bg=bg, fc=fc2, bold=True, ha='left', sz=11)
    for c in [6, 7, 8]: put(ws, r, c, '', bg=bg)
    ws.row_dimensions[r].height = 19

def spacer(ws, r):
    for c in range(1, 9): put(ws, r, c, '', bg=GREY)
    ws.row_dimensions[r].height = 6

# ── MAIN BUILDER ──────────────────────────────────────────────────────────────
def build_scorecard(xl_path, label, initial_cap):
    # ── LOAD SUMMARY ─────────────────────────────────────────────────────────
    summary_all = pd.read_excel(xl_path, sheet_name='Current_Summary')
    summary_all.columns = ['Date', 'Holdings', 'Invested', 'HoldValue',
                            'Cash', 'TotalValue', 'Diff', 'PnL', 'SIP', 'AddCap']
    summary_all = summary_all.dropna(subset=['Date'])
    summary_all['Date'] = pd.to_datetime(summary_all['Date'], dayfirst=True)

    daily = (summary_all.groupby('Date')
             .agg(SIPAdded=('SIP', 'sum'), AddCapAdded=('AddCap', 'sum'),
                  LastValue=('TotalValue', 'last'), LastPnL=('PnL', 'last'),
                  LastInvested=('Invested', 'last'))
             .reset_index().sort_values('Date'))

    latest         = daily.iloc[-1]
    strategy_start = daily.iloc[0]['Date']
    total_invested = float(latest['LastInvested'])
    current_value  = float(latest['LastValue'])
    current_pnl    = float(latest['LastPnL'])
    days_active    = (latest['Date'] - strategy_start).days
    total_sip      = daily['SIPAdded'].sum()
    total_add      = daily['AddCapAdded'].sum()
    simple_roic    = round(current_pnl / total_invested * 100, 2)
    annualised     = round(simple_roic * 365 / days_active, 2) if days_active > 0 else 0

    # XIRR
    cfs = [(strategy_start.date(), -initial_cap)]
    for _, row in daily.iterrows():
        if float(row['AddCapAdded']) > 1: cfs.append((row['Date'].date(), -float(row['AddCapAdded'])))
        if float(row['SIPAdded'])   > 1: cfs.append((row['Date'].date(), -float(row['SIPAdded'])))
    cfs.append((latest['Date'].date(), current_value))
    xirr_val = xirr(sorted(cfs, key=lambda x: x[0]))
    xirr_str = f'{xirr_val}% p.a.' if xirr_val is not None else 'N/A'
    xirr_fc  = GREEN if (xirr_val or 0) >= 0 else RED

    # ── SHARPE RATIO (6% RF, annualised from daily returns) ───────────────────
    daily_rets = daily['LastValue'].pct_change().dropna()
    rf_daily   = 0.06 / 252
    if len(daily_rets) > 1 and daily_rets.std() > 0:
        sharpe     = round((daily_rets.mean() - rf_daily) / daily_rets.std() * np.sqrt(252), 2)
        sharpe_str = str(sharpe)
        sharpe_fc  = GREEN if sharpe > 1 else (AMBER if sharpe > 0 else RED)
    else:
        sharpe, sharpe_str, sharpe_fc = None, 'N/A', '1F3864'

    # ── DRAWDOWN ──────────────────────────────────────────────────────────────
    daily['Peak']   = daily['LastValue'].cummax()
    daily['DD_pct'] = ((daily['LastValue'] - daily['Peak']) / daily['Peak'] * 100).round(2)
    peak_val  = daily['LastValue'].max()
    peak_date = daily.loc[daily['LastValue'].idxmax(), 'Date'].strftime('%d %b %Y')
    max_dd    = float(daily['DD_pct'].min())
    max_dd_dt = daily.loc[daily['DD_pct'].idxmin(), 'Date'].strftime('%d %b %Y')
    curr_dd   = round(((current_value - peak_val) / peak_val) * 100, 2)

    # Drawdown duration — longest consecutive stretch below peak
    in_dd_list = (daily['DD_pct'] < 0).tolist()
    max_dd_dur = cur_streak = 0
    for v in in_dd_list:
        cur_streak = cur_streak + 1 if v else 0
        max_dd_dur = max(max_dd_dur, cur_streak)
    curr_dd_dur = 0
    for v in reversed(in_dd_list):
        if v: curr_dd_dur += 1
        else: break

    # ── CALMAR RATIO ──────────────────────────────────────────────────────────
    if xirr_val is not None and xirr_val > 0 and max_dd != 0:
        calmar     = round(xirr_val / abs(max_dd), 2)
        calmar_str = str(calmar)
        calmar_fc  = GREEN if calmar > 1 else (AMBER if calmar > 0 else RED)
    else:
        calmar, calmar_str, calmar_fc = None, 'N/A', '1F3864'

    # ── PORTFOLIO ─────────────────────────────────────────────────────────────
    port_raw = pd.read_excel(xl_path, sheet_name='Current_Portfolio', header=None)
    port_stocks = []
    for i in range(1, len(port_raw)):
        row = port_raw.iloc[i]
        if pd.notna(row[6]):
            port_stocks.append({
                'Stock': str(row[6]), 'Holding_Days': int(row[8] or 0),
                'Buy_Amount': float(row[11] or 0), 'Curr_Amount': float(row[13] or 0),
                'PnL': float(row[15] or 0), 'Pct': float(row[16] or 0),
            })
    n_holdings     = int(port_raw.iloc[1, 1]) if pd.notna(port_raw.iloc[1, 1]) else 0
    cash_remaining = float(port_raw.iloc[1, 2]) if pd.notna(port_raw.iloc[1, 2]) else 0
    open_pnl       = sum(s['PnL'] for s in port_stocks)
    open_invested  = sum(s['Buy_Amount'] for s in port_stocks)
    open_pnl_pct   = round(open_pnl / open_invested * 100, 2) if open_invested else 0
    open_green     = sum(1 for s in port_stocks if s['PnL'] > 0)
    open_red       = sum(1 for s in port_stocks if s['PnL'] < 0)
    open_flat      = sum(1 for s in port_stocks if s['PnL'] == 0)

    # ── EXITS ─────────────────────────────────────────────────────────────────
    has_exits = False
    exited    = pd.DataFrame()
    try:
        exited = pd.read_excel(xl_path, sheet_name='Exited_Stocks')
        exited.columns = ['Stock', 'Entry_Date', 'Exit_Date', 'Holding_Days',
                          'Shares', 'Buy_Price', 'Buy_Amount', 'Sell_Price',
                          'Sell_Amount', 'EMA', 'Profit_Loss', 'Percentage', 'Exit_Type']
        has_exits = len(exited) > 0
    except Exception:
        pass

    # ── STOCK REPEAT ANALYSIS ─────────────────────────────────────────────────
    stock_repeat = pd.DataFrame()
    if has_exits:
        open_stocks = {s['Stock'] for s in port_stocks}
        sr = (exited.groupby('Stock')
              .agg(Rounds=('Profit_Loss', 'count'),
                   Wins=('Profit_Loss', lambda x: (x > 0).sum()),
                   TotalPnL=('Profit_Loss', 'sum'),
                   TotalInvested=('Buy_Amount', 'sum'),
                   AvgPct=('Percentage', 'mean'),
                   AvgDays=('Holding_Days', 'mean'))
              .reset_index())
        sr['Losses']    = sr['Rounds'] - sr['Wins']
        sr['NetRetPct'] = (sr['TotalPnL'] / sr['TotalInvested'] * 100).round(2)
        sr['Open']      = sr['Stock'].isin(open_stocks)
        sr = sr.sort_values(['Rounds', 'TotalPnL'], ascending=[False, False]).reset_index(drop=True)
        stock_repeat = sr

    # ── MONTHLY + CONSISTENCY ─────────────────────────────────────────────────
    daily['Month'] = daily['Date'].dt.to_period('M')
    monthly = (daily.groupby('Month')
               .agg(LastValue=('LastValue', 'last'), LastPnL=('LastPnL', 'last'),
                    SIPAdded=('SIPAdded', 'sum'), AddCap=('AddCapAdded', 'sum'))
               .reset_index())
    monthly['Monthly_PnL'] = monthly['LastPnL'].diff()
    monthly.loc[monthly.index[0], 'Monthly_PnL'] = monthly.iloc[0]['LastPnL']
    months_green = int((monthly['Monthly_PnL'] > 0).sum())
    total_months = len(monthly)

    # ── OPEN EXCEL ────────────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(xl_path)
    if 'Scorecard' in wb.sheetnames:
        del wb['Scorecard']
    ws = wb.create_sheet('Scorecard', 0)
    today = dt.datetime.now()

    put(ws, 1, 1, f'STRATEGY SCORECARD  \u2014  {label}',
        bg=DARK_BLUE, fc=WHITE, bold=True, sz=15, ha='center', merge_to=8, row_h=34)
    put(ws, 2, 1,
        f'Generated {today.strftime("%d %b %Y %H:%M")}  |  '
        f'{strategy_start.strftime("%d %b %Y")}  \u2192  {latest["Date"].strftime("%d %b %Y")}  ({days_active} days)',
        bg=MID_BLUE, fc=WHITE, italic=True, sz=9, ha='center', merge_to=8, row_h=18)
    spacer(ws, 3)

    r = 4
    # S1 — RETURNS ─────────────────────────────────────────────────────────────
    sec(ws, r, 'RETURNS'); r += 1
    pair(ws, r, 'Initial Capital', f'{RS}{initial_cap:,.0f}',
               'Total Invested (Cumulative)', f'{RS}{total_invested:,.0f}'); r += 1
    pair(ws, r, '  of which SIP', f'{RS}{total_sip:,.0f}',
               '  of which Lump Sum Added', f'{RS}{total_add:,.0f}'); r += 1
    pair(ws, r, 'Current Portfolio Value', f'{RS}{current_value:,.0f}',
               'Net P&L (mark-to-market)', f'{RS}{current_pnl:,.0f}',
               fc2=GREEN if current_pnl >= 0 else RED); r += 1
    pair(ws, r, 'ROIC (on total invested)', f'{simple_roic}%',
               'Annualised Return (simple)', f'{annualised}% p.a.',
               fc1=GREEN if simple_roic >= 0 else RED,
               fc2=GREEN if annualised >= 0 else RED); r += 1
    pair(ws, r, 'XIRR', xirr_str,
               'Days Active', str(days_active), fc1=xirr_fc); r += 1
    spacer(ws, r); r += 1

    # S2 — MONTHLY P&L + CONSISTENCY SCORE ────────────────────────────────────
    sec(ws, r, f'MONTHLY P&L  \u2014  Consistency: {months_green} of {total_months} months green'); r += 1
    for col, hdr in enumerate(['Month', f'End Value ({RS})', f'M-o-M P&L ({RS})', f'SIP ({RS})', f'Lump Sum ({RS})', '', '', ''], 1):
        put(ws, r, col, hdr, bg=LT_BLUE, bold=True, ha='center', sz=10)
    ws.row_dimensions[r].height = 18; r += 1
    for _, row in monthly.iterrows():
        bg  = VLT_BLUE if r % 2 == 0 else WHITE
        pfc = GREEN if row['Monthly_PnL'] >= 0 else RED
        put(ws, r, 1, str(row['Month']),            bg=bg, bold=True, sz=10)
        put(ws, r, 2, round(row['LastValue'], 0),   bg=bg, ha='right', sz=10)
        put(ws, r, 3, round(row['Monthly_PnL'], 0), bg=bg, ha='right', fc=pfc, bold=True, sz=10)
        put(ws, r, 4, round(row['SIPAdded'], 0),    bg=bg, ha='right', sz=10)
        put(ws, r, 5, round(row['AddCap'], 0),      bg=bg, ha='right', sz=10)
        for c in [6, 7, 8]: put(ws, r, c, '', bg=bg)
        ws.row_dimensions[r].height = 17; r += 1
    spacer(ws, r); r += 1

    # S3 — STOCK REPEAT ANALYSIS ──────────────────────────────────────────────
    if has_exits and len(stock_repeat) > 0:
        sec(ws, r, f'STOCK REPEAT ANALYSIS  ({len(stock_repeat)} stocks traded)'); r += 1
        for col, hdr in enumerate(['Stock', 'Rounds', 'W / L', f'Net PnL ({RS})',
                                    f'Invested ({RS})', 'Net Ret %', 'Avg Days', 'Open?'], 1):
            put(ws, r, col, hdr, bg=LT_BLUE, bold=True, ha='center', sz=10)
        ws.row_dimensions[r].height = 18; r += 1

        for _, srow in stock_repeat.iterrows():
            bg   = VLT_BLUE if r % 2 == 0 else WHITE
            pf   = GREEN if srow['TotalPnL'] >= 0 else RED
            rf   = GREEN if srow['NetRetPct'] >= 0 else RED
            wlfc = GREEN if srow['Wins'] > srow['Losses'] else (RED if srow['Losses'] > srow['Wins'] else AMBER)
            name = srow['Stock'].split(':')[-1]
            put(ws, r, 1, name,                                        bg=bg, bold=True, sz=10)
            put(ws, r, 2, int(srow['Rounds']),                         bg=bg, ha='center', sz=10)
            put(ws, r, 3, f"{int(srow['Wins'])}W / {int(srow['Losses'])}L", bg=bg, ha='center', fc=wlfc, sz=10)
            put(ws, r, 4, round(srow['TotalPnL'], 0),                  bg=bg, ha='right', fc=pf, bold=True, sz=10)
            put(ws, r, 5, round(srow['TotalInvested'], 0),             bg=bg, ha='right', sz=10)
            put(ws, r, 6, f"{srow['NetRetPct']:+.2f}%",                bg=bg, ha='center', fc=rf, bold=True, sz=10)
            put(ws, r, 7, f"{srow['AvgDays']:.0f}d",                   bg=bg, ha='center', sz=10)
            put(ws, r, 8, 'YES' if srow['Open'] else '',               bg=bg, ha='center',
                fc=GREEN if srow['Open'] else '000000', bold=bool(srow['Open']), sz=10)
            ws.row_dimensions[r].height = 17; r += 1
        spacer(ws, r); r += 1

    # S4 — RISK-ADJUSTED PERFORMANCE ──────────────────────────────────────────
    sec(ws, r, 'RISK-ADJUSTED PERFORMANCE'); r += 1
    pair(ws, r, 'Sharpe Ratio  (6% RF, annualised)', sharpe_str,
               'Calmar Ratio  (XIRR \u00f7 |Max DD|)', calmar_str,
               fc1=sharpe_fc, fc2=calmar_fc); r += 1
    pair(ws, r, 'Peak Portfolio Value', f'{RS}{peak_val:,.0f}  ({peak_date})',
               'Max Drawdown', f'{max_dd:.2f}%  ({max_dd_dt})',
               fc2=RED if max_dd < -5 else AMBER); r += 1
    pair(ws, r, 'Max Drawdown Duration', f'{max_dd_dur} days below peak',
               'Current vs Peak', f'{curr_dd:.2f}%',
               fc1=RED if max_dd_dur > 60 else (AMBER if max_dd_dur > 30 else '1F3864'),
               fc2=RED if curr_dd < -5 else (AMBER if curr_dd < 0 else GREEN)); r += 1
    if curr_dd_dur > 0:
        pair(ws, r, 'Currently in Drawdown', f'{curr_dd_dur} days and counting', '', '',
             fc1=RED if curr_dd_dur > 30 else AMBER); r += 1
    spacer(ws, r); r += 1

    # S3 — TRADE QUALITY ───────────────────────────────────────────────────────
    if has_exits:
        n_total  = len(exited)
        n_wins   = int((exited['Profit_Loss'] > 0).sum())
        n_losses = n_total - n_wins
        win_rate = round(n_wins / n_total * 100, 1)
        winners  = exited[exited['Profit_Loss'] > 0]
        losers   = exited[exited['Profit_Loss'] <= 0]
        avg_win_pct   = round(winners['Percentage'].mean(), 2) if len(winners) else 0
        avg_loss_pct  = round(losers['Percentage'].mean(), 2)  if len(losers)  else 0
        avg_win_days  = round(winners['Holding_Days'].mean(), 1) if len(winners) else 0
        avg_loss_days = round(losers['Holding_Days'].mean(), 1)  if len(losers)  else 0
        best      = exited.loc[exited['Percentage'].idxmax()]
        worst     = exited.loc[exited['Percentage'].idxmin()]
        net_closed = exited['Profit_Loss'].sum()
        comb       = net_closed + open_pnl

        sec(ws, r, f'TRADE QUALITY  ({n_total} exits)'); r += 1
        pair(ws, r, 'Win Rate', f'{win_rate}%  ({n_wins}W / {n_losses}L)',
                   'Net PnL \u2014 Closed Trades', f'{RS}{net_closed:,.0f}',
                   fc1=GREEN if win_rate >= 50 else (AMBER if win_rate >= 30 else RED),
                   fc2=GREEN if net_closed >= 0 else RED); r += 1
        pair(ws, r, 'Avg Gain (winners)', f'+{avg_win_pct}%  avg {avg_win_days}d held',
                   'Avg Loss (losers)', f'{avg_loss_pct}%  avg {avg_loss_days}d held',
                   fc1=GREEN, fc2=RED); r += 1
        pair(ws, r,
             'Best Exit',  f'{best["Stock"].split(":")[-1]}  +{best["Percentage"]:.2f}%  ({int(best["Holding_Days"])}d)',
             'Worst Exit', f'{worst["Stock"].split(":")[-1]}  {worst["Percentage"]:.2f}%  ({int(worst["Holding_Days"])}d)',
             fc1=GREEN, fc2=RED); r += 1
        pair(ws, r, 'Net PnL open positions', f'{RS}{open_pnl:,.0f}',
                   'Combined PnL (closed + open)', f'{RS}{comb:,.0f}',
                   fc1=GREEN if open_pnl >= 0 else RED,
                   fc2=GREEN if comb >= 0 else RED); r += 1
        spacer(ws, r); r += 1

        # S4 — EXIT TYPE BREAKDOWN ─────────────────────────────────────────────
        by_type = (exited.groupby('Exit_Type')
                   .agg(Count=('Profit_Loss', 'count'), NetPnL=('Profit_Loss', 'sum'),
                        AvgPct=('Percentage', 'mean'), AvgDays=('Holding_Days', 'mean'))
                   .reset_index())
        sec(ws, r, 'EXIT TYPE BREAKDOWN'); r += 1
        for col, hdr in enumerate(['Exit Type', 'Count', f'Net PnL ({RS})', 'Avg %', 'Avg Days', '', '', ''], 1):
            put(ws, r, col, hdr, bg=LT_BLUE, bold=True, ha='center', sz=10)
        ws.row_dimensions[r].height = 18; r += 1
        for _, row in by_type.iterrows():
            bg = VLT_BLUE if r % 2 == 0 else WHITE
            nf = GREEN if row['NetPnL'] >= 0 else RED
            af = GREEN if row['AvgPct'] >= 0 else RED
            put(ws, r, 1, str(row['Exit_Type']),        bg=bg, bold=True, sz=10)
            put(ws, r, 2, int(row['Count']),             bg=bg, ha='center', sz=10)
            put(ws, r, 3, round(row['NetPnL'], 0),       bg=bg, ha='right', fc=nf, bold=True, sz=10)
            put(ws, r, 4, f"{row['AvgPct']:+.1f}%",      bg=bg, ha='center', fc=af, sz=10)
            put(ws, r, 5, f"{row['AvgDays']:.0f}d",      bg=bg, ha='center', sz=10)
            for c in [6, 7, 8]: put(ws, r, c, '', bg=bg)
            ws.row_dimensions[r].height = 17; r += 1
        spacer(ws, r); r += 1

        # S5 — EMA BREACH LOSS BUCKETS ─────────────────────────────────────────
        ema_exits  = exited[exited['Exit_Type'].str.contains('EMA', case=False, na=False)]
        ema_losses = ema_exits[ema_exits['Percentage'] < 0]
        if len(ema_losses) > 0:
            n_ema = len(ema_losses)
            b_lt5  = ema_losses[ema_losses['Percentage'].abs() < 5]
            b_5_10 = ema_losses[(ema_losses['Percentage'].abs() >= 5) & (ema_losses['Percentage'].abs() < 10)]
            b_gt10 = ema_losses[ema_losses['Percentage'].abs() >= 10]

            def bkt(df):
                if len(df) == 0: return '0 exits'
                return f'{len(df)} exits  |  {RS}{df["Profit_Loss"].sum():,.0f}  |  avg {df["Percentage"].mean():+.1f}%'

            note_lt5  = 'Cutting on noise'     if len(b_lt5)  > n_ema * 0.5 else ''
            note_gt10 = 'EMA reacting slowly'  if len(b_gt10) > n_ema * 0.3 else ''

            sec(ws, r, f'EMA BREACH LOSS BUCKETS  (exit-speed diagnostic, {n_ema} loss exits)'); r += 1
            pair(ws, r, 'Loss < 5%   (tight cuts)',   bkt(b_lt5),  '', note_lt5,
                 fc1=GREEN if len(b_lt5) > 0 else '1F3864'); r += 1
            pair(ws, r, 'Loss 5\u201310%  (normal zone)',  bkt(b_5_10), '', ''); r += 1
            pair(ws, r, 'Loss > 10%  (slow to react)', bkt(b_gt10), '', note_gt10,
                 fc1=RED if len(b_gt10) > 0 else GREEN); r += 1
            spacer(ws, r); r += 1
    else:
        sec(ws, r, 'TRADE QUALITY'); r += 1
        pair(ws, r, 'Exits', 'No exits recorded yet', '', ''); r += 1
        spacer(ws, r); r += 1

    # S6 — OPEN POSITIONS ──────────────────────────────────────────────────────
    sec(ws, r, 'OPEN POSITIONS'); r += 1
    pair(ws, r, 'Holdings / Cash', f'{n_holdings} stocks  |  {RS}{cash_remaining:,.0f} cash', '', ''); r += 1
    pair(ws, r, 'Open Invested', f'{RS}{open_invested:,.0f}',
               'Unrealized PnL', f'{RS}{open_pnl:,.0f}  ({open_pnl_pct}%)',
               fc2=GREEN if open_pnl >= 0 else RED); r += 1
    pair(ws, r, 'Green / Flat / Red', f'{open_green}G / {open_flat}F / {open_red}R', '', ''); r += 1
    spacer(ws, r); r += 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 26
    ws.column_dimensions['F'].width = 11
    ws.column_dimensions['G'].width = 11
    ws.column_dimensions['H'].width = 9
    ws.freeze_panes = 'A4'

    wb.save(xl_path)
    print(f'Scorecard written \u2192 {xl_path}')


if __name__ == '__main__':
    import io
    from pathlib import Path
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    os.chdir(Path(__file__).parent.parent)
    build_scorecard(
        'momentum/ZU9940_old/Logs_EMA/PRAMOD_Z_Consolidated_Momentum_Results.xlsx',
        'MOMENTUM STOCKS \u2014 ZU9940 (Jun 2025 \u2013 Dec 2025)',
        initial_cap=1_000_000,
    )
    build_scorecard(
        'momentum/ZU9940/Logs_EMA/PRAMOD_Z_Consolidated_Momentum_Results.xlsx',
        'MOMENTUM STOCKS \u2014 ZU9940 (Dec 2025 \u2013 Present)',
        initial_cap=2_500_000,
    )
    build_scorecard(
        'momentum_etf/ZU9940/Logs_ETF/PRAMOD_Z_Consolidated_Momentum_Results.xlsx',
        'MOMENTUM ETF \u2014 ZU9940',
        initial_cap=200_000,
    )

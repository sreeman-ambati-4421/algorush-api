"""
One-off backfill: reads the existing per-account Excel workbooks produced by
the OLD EquityAlgo repo's momentum_final.py / momentum_etf.py and loads their
history into Postgres (Neon), using the same schema AlgoRush-API's bot now
writes to directly (see db/schema.sql and Algo/utils/db.py in THIS repo).

Run this ONCE per account/strategy, on the host that has the real workbooks
(EC2 or wherever the old bot has been running), after applying db/schema.sql
to your Neon database and setting DATABASE_URL. Always imports Algo.utils.db
/ Algo.utils.creds from THIS repo (AlgoRush-API) -- --xlsx-source-root is only
used as a plain filesystem path to locate the old .xlsx files, never added to
sys.path, so there's no collision with the old repo's own Algo package.

Usage:
    # Apply schema.sql once:
    psql "$DATABASE_URL" -f db/schema.sql

    # Auto-discover every account in ACCOUNTS (from this repo's Algo/utils/creds.py)
    # and migrate whatever workbook exists for each account/strategy combination,
    # reading the old repo's on-disk layout (<root>/Algo/momentum/<account>/Logs_EMA/...):
    python db/migrate_excel_to_db.py --xlsx-source-root /path/to/old/EquityAlgo

    # Or migrate a single workbook explicitly:
    python db/migrate_excel_to_db.py \\
        --account ZU9940 --strategy momentum \\
        --xlsx /path/to/old/EquityAlgo/Algo/momentum/ZU9940/Logs_EMA/PRAMOD_Z_Consolidated_Momentum_Results.xlsx \\
        --client-name PRAMOD_Z --initial-capital 2500000
"""

import argparse
import os
import sys
from collections import OrderedDict

import pandas as pd
from openpyxl import load_workbook

# Make this repo's Algo package importable regardless of cwd.
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

STRATEGY_LOG_DIR = {"momentum": "Logs_EMA", "momentum_etf": "Logs_ETF"}
STRATEGY_SUBDIR = {"momentum": "Algo/momentum", "momentum_etf": "Algo/momentum_etf"}
STRATEGY_DEFAULT_CAPITAL = {"momentum": 500000, "momentum_etf": 100000}
STRATEGY_DEFAULT_REBAL = {"momentum": 5, "momentum_etf": 5}


def _read_current_portfolio(xlsx_path):
    wb = load_workbook(filename=xlsx_path, data_only=True)
    if "Current_Portfolio" not in wb.sheetnames:
        return None
    rows = list(wb["Current_Portfolio"].iter_rows(values_only=True))
    if len(rows) < 2:
        return None
    data = {
        "Executed_Date": rows[1][0],
        "No_of_Holdings": rows[1][1] or 0,
        "Cash_Remaining": rows[1][2] or 0,
        "Rebalance_Counter": rows[1][3] or 1,
        "portfolio": OrderedDict(),
    }
    for row in rows[1:]:
        if not row[6]:
            continue
        data["portfolio"][row[6]] = {
            "Entry_Date": row[7],
            "Holding_Days": row[8] or 1,
            "No_Of_Shares": row[9],
            "Buy_Price": row[10],
            "Buy_Amount": row[11],
            "Current_Price": row[12] or 0,
            "Current_Amount": row[13] or 0,
            "100_Days_EMA": row[14] if len(row) > 14 else 0,
            "Profit_Loss": row[15] if len(row) > 15 else 0,
            "Percentage": row[16] if len(row) > 16 else 0,
        }
    return data


def _read_summary_history(xlsx_path):
    wb = load_workbook(filename=xlsx_path, data_only=True)
    if "Current_Summary" not in wb.sheetnames:
        return []
    keys = ("Date", "No_of_Holdings", "Invested_Capital", "Holdings_Value", "Cash_Remaining",
            "Total_Value_Holdings", "Holding_Values_Diff", "Total_Profit_Loss",
            "SIP_Amount_Added", "Additional_Capital_Added")
    out = []
    for row in wb["Current_Summary"].iter_rows(min_row=1, values_only=True):
        if not row or row[0] is None or row[0] == "Date":
            continue
        out.append(dict(zip(keys, row)))
    return out


def _read_exited_stocks(xlsx_path):
    try:
        df = pd.read_excel(xlsx_path, sheet_name="Exited_Stocks")
    except ValueError:
        return []
    records = []
    for _, row in df.iterrows():
        records.append({
            "ticker": row["Exit Stock"],
            "Entry_Date": row["Entry_Date"],
            "Exit_Date": row["Exit_Date"],
            "Holding_Days": int(row["Holding_Days"]),
            "No_Of_Shares": row["No_Of_Shares"],
            "Buy_Price": row["Buy_Price"],
            "Buy_Amount": row["Buy_Amount"],
            "Sell_Price": row["Sell_Price"],
            "Sell_Amount": row["Sell_Amount"],
            "100_Days_EMA": row.get("100_Days_EMA", 0),
            "Profit_Loss": row["Profit_Loss"],
            "Percentage": row["Percentage"],
            "Exit_Type": row.get("Exit_Type", "UNKNOWN"),
        })
    return records


def migrate_one(db, account_id, strategy, client_name, xlsx_path, initial_capital):
    if not os.path.exists(xlsx_path):
        print(f"  [skip] {account_id}/{strategy}: no workbook at {xlsx_path}")
        return

    print(f"  [migrate] {account_id}/{strategy}: {xlsx_path}")
    db.ensure_account(account_id, client_name)

    settings = db.load_strategy_settings(strategy)
    if account_id not in settings:
        db.save_strategy_settings(strategy, {
            account_id: {
                "reset_and_rebalance": False,
                "rebalance_today": False,
                "initial_capital": initial_capital,
                "sip_amount": 0,
                "additional_capital": 0,
                "rebalance_counter": STRATEGY_DEFAULT_REBAL[strategy],
            }
        })
        print(f"    seeded strategy_settings with initial_capital={initial_capital} "
              f"-- verify this is correct, it could not be inferred from the workbook")

    summary_rows = _read_summary_history(xlsx_path)
    for row in summary_rows:
        date_val = row["Date"]
        date_str = date_val if isinstance(date_val, str) else date_val.strftime("%d-%m-%Y")
        db.write_current_summary(account_id, strategy, {
            "Date": date_str,
            "No_of_Holdings": row["No_of_Holdings"] or 0,
            "Invested_Capital": row["Invested_Capital"] or 0,
            "Holdings_Value": row["Holdings_Value"] or 0,
            "Cash_Remaining": row["Cash_Remaining"] or 0,
            "Total_Value_Holdings": row["Total_Value_Holdings"] or 0,
            "Holding_Values_Diff": row.get("Holding_Values_Diff"),
            "Total_Profit_Loss": row["Total_Profit_Loss"] or 0,
            "SIP_Amount_Added": row.get("SIP_Amount_Added") or 0,
            "Additional_Capital_Added": row.get("Additional_Capital_Added") or 0,
        })
    print(f"    summary_history: {len(summary_rows)} row(s)")

    exits = _read_exited_stocks(xlsx_path)
    # write_exit_stocks() takes a dict keyed by ticker, so batching more than
    # one row per call silently drops earlier rows whenever the same ticker
    # was exited more than once under the same exit_type (a real occurrence
    # in multi-year history). Write one row per call instead -- each call's
    # dict has exactly one entry, so no ticker can collide with another.
    for r in exits:
        entry_date = r["Entry_Date"] if isinstance(r["Entry_Date"], str) else r["Entry_Date"].strftime("%d-%m-%Y")
        exit_date = r["Exit_Date"] if isinstance(r["Exit_Date"], str) else r["Exit_Date"].strftime("%d-%m-%Y")
        db.write_exit_stocks(account_id, strategy, {
            r["ticker"]: {
                "Entry_Date": entry_date, "Exit_Date": exit_date,
                "Holding_Days": r["Holding_Days"], "No_Of_Shares": r["No_Of_Shares"],
                "Buy_Price": r["Buy_Price"], "Buy_Amount": r["Buy_Amount"],
                "Sell_Price": r["Sell_Price"], "Sell_Amount": r["Sell_Amount"],
                "100_Days_EMA": r["100_Days_EMA"], "Profit_Loss": r["Profit_Loss"],
                "Percentage": r["Percentage"],
            }
        }, r["Exit_Type"])
    print(f"    exited_stocks: {len(exits)} row(s)")

    portfolio = _read_current_portfolio(xlsx_path)
    if portfolio:
        exec_date = portfolio["Executed_Date"]
        portfolio["Executed_Date"] = exec_date if isinstance(exec_date, str) else exec_date.strftime("%d-%m-%Y")
        for ticker, d in portfolio["portfolio"].items():
            ed = d["Entry_Date"]
            d["Entry_Date"] = ed if isinstance(ed, str) else ed.strftime("%d-%m-%Y")
        db.write_current_portfolio(account_id, strategy, portfolio)
        print(f"    portfolio_holdings: {len(portfolio['portfolio'])} open position(s)")


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--xlsx-source-root", help="Path to the OLD repo root that has the real .xlsx workbooks (for auto-discover)")
    parser.add_argument("--account", help="Single account userid to migrate")
    parser.add_argument("--strategy", choices=["momentum", "momentum_etf"], help="Strategy for --account")
    parser.add_argument("--xlsx", help="Path to the workbook for --account/--strategy")
    parser.add_argument("--client-name", help="Client name for --account (used in ensure_account)")
    parser.add_argument("--initial-capital", type=float, help="Seed value for strategy_settings.initial_capital")
    args = parser.parse_args()

    import Algo.utils.db as db  # noqa: E402
    from Algo.utils.creds import ACCOUNTS  # noqa: E402

    if args.account:
        if not (args.strategy and args.xlsx and args.client_name):
            parser.error("--account requires --strategy, --xlsx, and --client-name")
        migrate_one(
            db, args.account, args.strategy, args.client_name, args.xlsx,
            args.initial_capital or STRATEGY_DEFAULT_CAPITAL[args.strategy],
        )
        return

    if not args.xlsx_source_root:
        parser.error("pass --xlsx-source-root to auto-discover, or use --account/--strategy/--xlsx for a single workbook")

    for account_id, cfg in ACCOUNTS.items():
        client_name = cfg.get("client_name", account_id)
        for strategy, log_dir in STRATEGY_LOG_DIR.items():
            xlsx_path = os.path.join(
                args.xlsx_source_root, STRATEGY_SUBDIR[strategy], account_id, log_dir,
                f"{client_name}_Consolidated_Momentum_Results.xlsx",
            )
            migrate_one(db, account_id, strategy, client_name, xlsx_path, STRATEGY_DEFAULT_CAPITAL[strategy])


if __name__ == "__main__":
    main()
